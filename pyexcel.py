#!/usr/bin/python3

import openpyxl


class PyExcel:
    def __init__(self):
        self.__wb = None
        self.__sheet = None

    def __iter__(self):
        for x in self.__wb.sheetnames:
            yield x

    def IsLoaded(self):
        return self.__wb is not None

    def Reset(self):
        if not self.IsLoaded():
            print("not loaded.")
            return
        self.__wb = None

    def Save(self, fname):
        self.__wb.save(fname)

    def Load(self, fname=None):
        if self.IsLoaded():
            print("already load.")
            return
        if fname is None:
            # Create new Wb
            self.__wb = openpyxl.Workbook()
        else:
            self.__wb = openpyxl.load_workbook(fname)

    def GetSheetNames(self):
        if not self.IsLoaded():
            print("not loaded.")
            return
        return self.__wb.sheetnames

    def SetSheet(self, sheet_name):
        if sheet_name not in self.__wb.sheetnames:
            print("not have sheet %s" % sheet_name)
            return
        self.__sheet = PySheet(self.__wb[sheet_name])
        return self.__sheet

    def AddSheet(self, sheet_name, with_set=True):
        sheet = self.__wb.create_sheet(sheet_name)
        if with_set:
            self.__sheet = PySheet(sheet)
            return self.__sheet

    def CopySheet(self, sheet_name, with_set=True):
        sheet = self.__wb.copy_worksheet(self.__wb[sheet_name])
        if with_set:
            self.__sheet = PySheet(sheet)
            return self.__sheet

    def RemoveSheet(self, sheet_name):
        self.__wb.remove_sheet(self.__wb[sheet_name])


class PySheet:
    alphabet = [chr(i) for i in range(ord('a'), ord('z') + 1)]

    def __init__(self, sheet):
        self.__sheet = sheet

    def __none(self, *data):
        if type(data) == list:
            for d in data:
                if d is None:
                    return True
            return False
        else:
            return data is None

    def __num_to_col(self, d):
        res = ''
        alp_len = len(self.alphabet)
        while d != 0:
            res = (self.alphabet[d % alp_len]) + res
            d = int(d/alp_len)
        return res

    def __set_col_data(self, row, start_col, values):
        if hasattr(values, '__iter__'):
            for x, v in enumerate(values):
                self.SetValue(row, start_col+x, v)
        else:
            self.SetValue(row, start_col, values)

    def SetValue(self, cell_row, cell_col, value):
        self.__sheet.cell(row=cell_row, column=cell_col, value=value)

    def SetValues(self, start_row, start_col, values):
        if hasattr(values, '__iter__'):
            for y, row in enumerate(values):
                self.__set_col_data(start_row+y, start_col, row)
        else:
            self.SetValue(start_row, start_col, values)

    def __get_values(self, cell_range):
        return [[c.value for c in row] for row in self.__get_cells(cell_range)]

    def __get_cells(self, cell_range):
        if self.__sheet is None:
            print("not loaded sheet.")
            return
        return self.__sheet[cell_range]

    def GetValues(self, start_row=None, start_col=None, end_row=None, end_col=None, cell_range=None):
        if cell_range is not None:
            return self.__get_values(cell_range)
        elif self.__none(start_row, start_col, end_row, end_col):
            cell_range = '%s%d:%s%d' % (self.__num_to_col(
                start_col), start_row, self.__num_to_col(end_col), end_row)
            return self.__get_values(cell_range)
        else:
            print('not enough value')

    def GetSheetName(self):
        return self.__sheet.title

    def SetSheetName(self, title):
        self.__sheet.title = title
